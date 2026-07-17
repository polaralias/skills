#!/usr/bin/env python3
"""Generate daily and weekly Excel Gantt workbooks from a generic task list."""

from __future__ import annotations

import argparse
import json
from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any

try:
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ModuleNotFoundError as exc:  # pragma: no cover
    raise SystemExit(
        "Missing required Python dependency: "
        f"{exc.name}. Install the tasklist-gantt-creator requirements first with "
        "`python -m pip install -r requirements.txt` or run `scripts/bootstrap.ps1`."
    ) from exc


DEFAULT_CONFIG = {
    "input": {
        "path": "tasks.xlsx",
        "sheet": "Tasks",
        "skip_rows": 0,
        "format": "excel",
    },
    "columns": {
        "task_id": "Task ID",
        "task_name": "Task Name",
        "parent_id": "Parent ID",
        "parent_name": "Parent Name",
        "start_date": "Start Date",
        "due_date": "Due Date",
        "owner": "Owner",
        "customer": "Customer",
        "status": "Status",
    },
    "timeline": {
        "start": None,
        "end": None,
    },
    "output": {
        "path": "gantt.xlsx",
        "modes": ["daily", "weekly"],
        "focus_sheets": False,
        "colour_by": "owner",
        "omit_owner": False,
        "ignore_parent_bars": False,
    },
    "colours": {
        "owner": {},
        "status": {},
        "neutral": "BFBFBF",
        "header_dark": "1F3864",
        "header_mid": "2E5090",
        "header_light": "3A6BC5",
        "parent_bar": "5B9BD5",
        "milestone_bar": "1F3864",
        "task_alt": "F7F9FC",
        "task_plain": "FFFFFF",
        "milestone_row": "F2F2F2",
        "week_grid": "D9E2F3",
        "text_muted": "44546A",
        "status_palette": [
            "5B8FF9",
            "5AD8A6",
            "5D7092",
            "F6BD16",
            "6F5EF9",
            "6DC8EC",
            "F6903D",
            "008685",
            "F08BB4",
            "7F7F7F",
        ],
    },
}


@dataclass
class Theme:
    header_dark: str
    header_mid: str
    header_light: str
    parent_bar: str
    milestone_bar: str
    task_alt: str
    task_plain: str
    milestone_row: str
    week_grid: str
    text_muted: str
    neutral: str
    owner_colours: dict[str, str]
    status_colours: dict[str, str]
    status_palette: list[str]


def clean_text(value: Any) -> str:
    if pd.isnull(value):
        return ""
    text = str(value).strip()
    return "" if text.lower() == "nan" else text


def safe_excel_text(value: Any) -> str:
    """Keep imported labels as text instead of executable spreadsheet formulas."""
    text = clean_text(value)
    raw_text = "" if pd.isnull(value) else str(value)
    if raw_text.startswith(("\t", "\r")) or text.startswith(("=", "+", "-", "@")):
        return f"'{text}"
    return text


def normalize_id(value: Any) -> str:
    text = clean_text(value)
    if text.endswith(".0"):
        return text[:-2]
    return text


def parse_date_string(value: str) -> datetime:
    return datetime.strptime(value, "%Y-%m-%d")


def load_json(path: Path) -> dict[str, Any]:
    with path.open("r", encoding="utf-8") as handle:
        return json.load(handle)


def merge_dict(base: dict[str, Any], override: dict[str, Any]) -> dict[str, Any]:
    merged = dict(base)
    for key, value in override.items():
        if isinstance(value, dict) and isinstance(merged.get(key), dict):
            merged[key] = merge_dict(merged[key], value)
        else:
            merged[key] = value
    return merged


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Generate a Gantt workbook from a generic task list.")
    parser.add_argument("--config", required=True, help="Path to the JSON config file.")
    parser.add_argument("--input", help="Override input path.")
    parser.add_argument("--output", help="Override output path.")
    parser.add_argument("--start", help="Override timeline start date (YYYY-MM-DD).")
    parser.add_argument("--end", help="Override timeline end date (YYYY-MM-DD).")
    parser.add_argument("--sheet", help="Override worksheet name.")
    parser.add_argument("--skip-rows", type=int, help="Override skipped header rows.")
    return parser.parse_args()


def apply_overrides(config: dict[str, Any], args: argparse.Namespace) -> dict[str, Any]:
    updated = merge_dict({}, config)
    if args.input:
        updated["input"]["path"] = args.input
    if args.output:
        updated["output"]["path"] = args.output
    if args.start:
        updated["timeline"]["start"] = args.start
    if args.end:
        updated["timeline"]["end"] = args.end
    if args.sheet:
        updated["input"]["sheet"] = args.sheet
    if args.skip_rows is not None:
        updated["input"]["skip_rows"] = args.skip_rows
    return updated


def build_theme(config: dict[str, Any]) -> Theme:
    colours = config["colours"]
    return Theme(
        header_dark=colours["header_dark"],
        header_mid=colours["header_mid"],
        header_light=colours["header_light"],
        parent_bar=colours["parent_bar"],
        milestone_bar=colours["milestone_bar"],
        task_alt=colours["task_alt"],
        task_plain=colours["task_plain"],
        milestone_row=colours["milestone_row"],
        week_grid=colours["week_grid"],
        text_muted=colours["text_muted"],
        neutral=colours["neutral"],
        owner_colours=colours.get("owner", {}),
        status_colours=colours.get("status", {}),
        status_palette=colours.get("status_palette", []),
    )


def read_input(config: dict[str, Any]) -> pd.DataFrame:
    input_cfg = config["input"]
    path = Path(input_cfg["path"])
    file_format = input_cfg.get("format", "").lower() or path.suffix.lower().lstrip(".")
    skip_rows = input_cfg.get("skip_rows", 0)

    if file_format in {"xlsx", "xlsm", "xls", "excel"}:
        return pd.read_excel(path, sheet_name=input_cfg.get("sheet", "Tasks"), skiprows=skip_rows)
    if file_format == "csv":
        return pd.read_csv(path, skiprows=skip_rows)
    raise ValueError(f"Unsupported input format: {file_format}")


def normalize_frame(raw: pd.DataFrame, config: dict[str, Any]) -> pd.DataFrame:
    columns = config["columns"]
    required = ["task_id", "task_name", "parent_id", "start_date", "due_date"]
    missing = [key for key in required if columns.get(key) not in raw.columns]
    if missing:
        raise ValueError(f"Missing required mapped columns: {', '.join(missing)}")

    frame = pd.DataFrame(
        {
            "task_id": raw[columns["task_id"]].apply(normalize_id),
            "task_name": raw[columns["task_name"]].apply(clean_text),
            "parent_id": raw[columns["parent_id"]].apply(normalize_id),
            "parent_name": raw[columns["parent_name"]].apply(clean_text)
            if columns.get("parent_name") in raw.columns
            else "",
            "start_date": pd.to_datetime(raw[columns["start_date"]], errors="coerce"),
            "due_date": pd.to_datetime(raw[columns["due_date"]], errors="coerce"),
            "owner_raw": raw[columns["owner"]].apply(clean_text) if columns.get("owner") in raw.columns else "",
            "customer_raw": raw[columns["customer"]].apply(clean_text) if columns.get("customer") in raw.columns else "",
            "status_raw": raw[columns["status"]].apply(clean_text) if columns.get("status") in raw.columns else "",
        }
    )

    frame = frame[frame["task_id"].astype(bool) & frame["task_name"].astype(bool)].copy()
    frame["_source_order"] = range(len(frame))
    frame["owner"] = frame["owner_raw"].apply(choose_primary_label)
    frame["customer"] = frame["customer_raw"]
    frame["status"] = frame["status_raw"]
    return frame


def choose_primary_label(value: str) -> str:
    labels = [item.strip() for item in clean_text(value).split(",") if item.strip()]
    return labels[0] if labels else ""


def rebuild_hierarchy(frame: pd.DataFrame) -> pd.DataFrame:
    task_ids = set(frame["task_id"])
    children_map: dict[str, list[int]] = defaultdict(list)
    roots: list[int] = []

    for idx, row in frame.sort_values("_source_order").iterrows():
        parent_id = row["parent_id"]
        if parent_id and parent_id in task_ids:
            children_map[parent_id].append(idx)
        else:
            roots.append(idx)

    for child_indexes in children_map.values():
        child_indexes.sort(key=lambda item: frame.at[item, "_source_order"])

    ordered: list[tuple[int, int]] = []
    visited: set[int] = set()

    def visit(index: int, depth: int) -> None:
        if index in visited:
            return
        visited.add(index)
        ordered.append((index, depth))
        for child_index in children_map.get(frame.at[index, "task_id"], []):
            visit(child_index, depth + 1)

    for index in roots:
        visit(index, 0)
    for index in frame.sort_values("_source_order").index:
        visit(index, 0)

    ordered_frame = frame.loc[[idx for idx, _ in ordered]].copy()
    ordered_frame["depth"] = [depth for _, depth in ordered]
    child_parent_ids = set(ordered_frame["parent_id"])
    ordered_frame["has_children"] = ordered_frame["task_id"].isin(child_parent_ids)
    ordered_frame["has_parent"] = ordered_frame["parent_id"].astype(bool)
    ordered_frame["is_top_parent"] = ordered_frame["depth"] == 0
    ordered_frame["is_subtask_header"] = ordered_frame["has_parent"] & ordered_frame["has_children"]
    ordered_frame["is_parent"] = ordered_frame["is_top_parent"] | ordered_frame["is_subtask_header"]
    return ordered_frame


def compute_descendant_ranges(frame: pd.DataFrame) -> dict[str, tuple[pd.Timestamp, pd.Timestamp] | None]:
    task_ids = set(frame["task_id"])
    children_map: dict[str, list[int]] = defaultdict(list)

    for idx, row in frame.sort_values("_source_order").iterrows():
        parent_id = row["parent_id"]
        if parent_id and parent_id in task_ids:
            children_map[parent_id].append(idx)

    memo: dict[str, tuple[pd.Timestamp, pd.Timestamp] | None] = {}

    def visit(task_id: str) -> tuple[pd.Timestamp, pd.Timestamp] | None:
        if task_id in memo:
            return memo[task_id]

        starts: list[pd.Timestamp] = []
        ends: list[pd.Timestamp] = []
        for child_idx in children_map.get(task_id, []):
            child = frame.loc[child_idx]
            if pd.notnull(child["start_date"]):
                starts.append(child["start_date"])
            if pd.notnull(child["due_date"]):
                ends.append(child["due_date"])
            nested = visit(child["task_id"])
            if nested:
                nested_start, nested_end = nested
                if pd.notnull(nested_start):
                    starts.append(nested_start)
                if pd.notnull(nested_end):
                    ends.append(nested_end)

        if starts or ends:
            range_start = min(starts) if starts else min(ends)
            range_end = max(ends) if ends else max(starts)
            memo[task_id] = (range_start, range_end)
        else:
            memo[task_id] = None
        return memo[task_id]

    return {task_id: visit(task_id) for task_id in task_ids}


def infer_timeline(frame: pd.DataFrame, config: dict[str, Any]) -> tuple[datetime, datetime]:
    start_cfg = config["timeline"].get("start")
    end_cfg = config["timeline"].get("end")

    starts = frame["start_date"].dropna()
    ends = frame["due_date"].dropna()
    if start_cfg:
        start = parse_date_string(start_cfg)
    elif not starts.empty:
        start = starts.min().to_pydatetime()
    else:
        raise ValueError("Timeline start could not be inferred; provide timeline.start.")

    if end_cfg:
        end = parse_date_string(end_cfg)
    elif not ends.empty:
        end = ends.max().to_pydatetime()
    else:
        raise ValueError("Timeline end could not be inferred; provide timeline.end.")

    return start, end


def monday_floor(ts: pd.Timestamp | datetime) -> datetime:
    return ts - timedelta(days=ts.weekday())


def format_cell_date(ts: Any, with_year: bool = False) -> str:
    if pd.isnull(ts):
        return ""
    return ts.strftime("%d %b %Y" if with_year else "%d/%m/%y")


def solid(hex_rgb: str) -> PatternFill:
    return PatternFill(start_color=hex_rgb, end_color=hex_rgb, fill_type="solid")


def hex_to_rgb(hex_rgb: str) -> tuple[int, int, int]:
    value = hex_rgb.lstrip("#")
    if len(value) == 8:
        value = value[2:]
    return tuple(int(value[i : i + 2], 16) for i in (0, 2, 4))


def ideal_text_colour(hex_rgb: str) -> str:
    r, g, b = hex_to_rgb(hex_rgb)
    brightness = (r * 299 + g * 587 + b * 114) / 1000
    return "000000" if brightness > 150 else "FFFFFF"


def build_status_map(frame: pd.DataFrame, theme: Theme) -> dict[str, str]:
    statuses = [status for status in sorted(frame["status"].unique()) if status]
    generated = {}
    palette = theme.status_palette or [theme.neutral]
    for index, status in enumerate(statuses):
        if status not in theme.status_colours:
            generated[status] = palette[index % len(palette)]
    return {**generated, **theme.status_colours}


def bar_colour(row: pd.Series, colour_by: str, theme: Theme, status_map: dict[str, str]) -> str:
    if colour_by == "status":
        return status_map.get(row["status"], theme.neutral)
    if colour_by == "owner":
        return theme.owner_colours.get(row["owner"], theme.neutral) if row["owner"] else theme.neutral
    return theme.neutral


def make_border(theme: Theme) -> Border:
    side = Side(style="thin", color=theme.week_grid)
    return Border(left=side, right=side, top=side, bottom=side)


def style_main_headers(ws: Any, header_row: int, month_groups: list[tuple[int, int, str]], week_groups: list[tuple[int, int, str]], theme: Theme) -> None:
    border = make_border(theme)
    for start_col, end_col, label in month_groups:
        ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
        cell = ws.cell(1, start_col)
        cell.value = label
        cell.fill = solid(theme.header_dark)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
        for col in range(start_col, end_col + 1):
            ws.cell(1, col).border = border

    if header_row == 3:
        for start_col, end_col, label in week_groups:
            ws.merge_cells(start_row=2, start_column=start_col, end_row=2, end_column=end_col)
            cell = ws.cell(2, start_col)
            cell.value = label
            cell.fill = solid(theme.header_mid)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
            for col in range(start_col, end_col + 1):
                ws.cell(2, col).border = border

    for cell in ws[header_row]:
        cell.fill = solid(theme.header_mid)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border


def style_body_row(ws: Any, row_idx: int, row_type: str, header_cols_count: int, owner_col_idx: int | None, owner_colour: str | None, theme: Theme) -> None:
    border = make_border(theme)
    if row_type == "top_parent":
        fill = solid(theme.header_dark)
        font = Font(bold=True, color="FFFFFF")
    elif row_type == "sub_parent":
        fill = solid(theme.header_mid)
        font = Font(bold=True, color="FFFFFF")
    elif row_type == "milestone":
        fill = solid(theme.milestone_row)
        font = Font(color=theme.text_muted)
    else:
        base = theme.task_alt if row_idx % 2 else theme.task_plain
        fill = solid(base)
        font = Font(color="000000")

    for col_idx in range(1, header_cols_count + 1):
        cell = ws.cell(row_idx, col_idx)
        cell.fill = fill
        cell.font = font
        cell.border = border
        cell.alignment = Alignment(vertical="center", wrap_text=True)

    if owner_col_idx and owner_colour:
        owner_cell = ws.cell(row_idx, owner_col_idx)
        owner_cell.fill = solid(owner_colour)
        owner_cell.font = Font(color=ideal_text_colour(owner_colour), bold=True)
        owner_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        owner_cell.border = border


def stretch_parent_bar(ws: Any, row_idx: int, row: pd.Series, date_units: list[Any], date_start_col: int, descendant_ranges: dict[str, tuple[pd.Timestamp, pd.Timestamp] | None], fill_hex: str, header_cols_count: int, theme: Theme) -> None:
    fill = solid(fill_hex)
    for col in range(1, header_cols_count + 1):
        ws.cell(row_idx, col).fill = fill

    descendant_range = descendant_ranges.get(row["task_id"])
    start_value = row["start_date"]
    end_value = row["due_date"]
    if descendant_range:
        start_value = descendant_range[0] if pd.notnull(descendant_range[0]) else start_value
        end_value = descendant_range[1] if pd.notnull(descendant_range[1]) else end_value
    if pd.isnull(start_value) or pd.isnull(end_value):
        return

    for col, unit in enumerate(date_units, start=date_start_col):
        active = (
            start_value <= unit[1] and end_value >= unit[0]
            if isinstance(unit, tuple)
            else start_value <= unit <= end_value
        )
        if active:
            ws.cell(row_idx, col).fill = fill
            ws.cell(row_idx, col).border = make_border(theme)


def is_milestone_row(row: pd.Series) -> bool:
    name = clean_text(row["task_name"]).lower()
    start = row["start_date"]
    due = row["due_date"]
    return name.startswith("milestone") or (pd.notnull(start) and pd.notnull(due) and start == due)


def write_main_sheet(
    wb: Workbook,
    sheet_name: str,
    frame: pd.DataFrame,
    date_units: list[Any],
    date_labels: list[str],
    month_groups: list[tuple[int, int, str]],
    week_groups: list[tuple[int, int, str]],
    colour_by: str,
    omit_owner: bool,
    ignore_parent_bars: bool,
    descendant_ranges: dict[str, tuple[pd.Timestamp, pd.Timestamp] | None],
    theme: Theme,
) -> None:
    ws = wb.active if wb.active.max_row == 1 and wb.active["A1"].value is None else wb.create_sheet(sheet_name)
    ws.title = sheet_name
    ws.sheet_view.showGridLines = False

    header_cols = ["Task Name", "Start Date", "Due Date"]
    if not omit_owner:
        header_cols.append("Owner")
    header_cols_count = len(header_cols)
    date_start_col = header_cols_count + 1
    owner_col_idx = 4 if not omit_owner else None

    header_row = 3 if sheet_name == "Gantt - Daily" else 2
    ws.append(header_cols + date_labels)
    for _ in range(header_row - 1):
        ws.insert_rows(1)

    style_main_headers(ws, header_row, month_groups, week_groups, theme)
    status_map = build_status_map(frame, theme)

    for _, row in frame.iterrows():
        values = [safe_excel_text(row["task_name"]), format_cell_date(row["start_date"]), format_cell_date(row["due_date"])]
        if not omit_owner:
            values.append(safe_excel_text(row["owner"]) or None)
        ws.append(values)
        row_idx = ws.max_row
        row_colour = bar_colour(row, colour_by, theme, status_map)

        if row["is_top_parent"]:
            row_type = "top_parent"
        elif row["is_subtask_header"]:
            row_type = "sub_parent"
        elif is_milestone_row(row):
            row_type = "milestone"
        else:
            row_type = "task"

        style_body_row(ws, row_idx, row_type, header_cols_count, owner_col_idx, row_colour if owner_col_idx else None, theme)

        if not ignore_parent_bars and row["is_parent"]:
            stretch_parent_bar(
                ws,
                row_idx,
                row,
                date_units,
                date_start_col,
                descendant_ranges,
                theme.parent_bar if row["is_subtask_header"] else theme.header_dark,
                header_cols_count,
                theme,
            )
            continue

        if pd.notnull(row["start_date"]) and pd.notnull(row["due_date"]):
            for col, unit in enumerate(date_units, start=date_start_col):
                active = (
                    row["start_date"] <= unit[1] and row["due_date"] >= unit[0]
                    if isinstance(unit, tuple)
                    else row["start_date"] <= unit <= row["due_date"]
                )
                if active:
                    ws.cell(row_idx, col).fill = solid(row_colour)
                    ws.cell(row_idx, col).border = make_border(theme)

    if sheet_name == "Gantt - Daily":
        ws.freeze_panes = f"{get_column_letter(date_start_col)}4"
        widths = [66, 12, 12, 26]
        for idx, width in enumerate(widths[:header_cols_count], start=1):
            ws.column_dimensions[get_column_letter(idx)].width = width
        for col in range(date_start_col, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(col)].width = 4
    else:
        ws.freeze_panes = f"{get_column_letter(date_start_col)}3"
        widths = [63, 11, 11, 24]
        for idx, width in enumerate(widths[:header_cols_count], start=1):
            ws.column_dimensions[get_column_letter(idx)].width = width
        for col in range(date_start_col, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(col)].width = 12

    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 21
    if header_row == 3:
        ws.row_dimensions[3].height = 24
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(ws.max_column)}{header_row}"


def clean_sheet_title(title: str, existing_titles: set[str]) -> str:
    invalid = '[]:*?/\\'
    cleaned = "".join("-" if char in invalid else char for char in clean_text(title)) or "Sheet"
    cleaned = cleaned[:31]
    base = cleaned
    suffix = 2
    while cleaned in existing_titles:
        suffix_text = f" {suffix}"
        cleaned = f"{base[:31-len(suffix_text)]}{suffix_text}"
        suffix += 1
    existing_titles.add(cleaned)
    return cleaned


def focus_sheet_candidates(frame: pd.DataFrame) -> list[pd.Series]:
    candidates = []
    for _, row in frame.iterrows():
        if row["depth"] != 1 or not row["has_children"]:
            continue
        name = clean_text(row["task_name"]).lower()
        if name in {"milestones", "governance"}:
            continue
        candidates.append(row)
    return candidates


def descendants_for(frame: pd.DataFrame, parent_id: str) -> list[pd.Series]:
    selected = []
    selected_ids = {parent_id}
    for _, row in frame.iterrows():
        if row["task_id"] == parent_id:
            selected.append(row)
            continue
        if row["parent_id"] in selected_ids:
            selected.append(row)
            selected_ids.add(row["task_id"])
    return selected


def split_stage_activity(task_name: str) -> tuple[str, str]:
    text = clean_text(task_name)
    if ":" in text:
        stage, activity = text.split(":", 1)
        return stage.strip(), activity.strip()
    if text.lower().startswith("milestone"):
        activity = text.split(":", 1)[1].strip() if ":" in text else text
        return "Milestone", activity
    return "Task", text


def build_focus_sheet(wb: Workbook, title: str, rows: list[pd.Series], colour_by: str, theme: Theme) -> None:
    existing_titles = set(wb.sheetnames)
    ws = wb.create_sheet(clean_sheet_title(title, existing_titles))
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "E4"

    feature_row = rows[0]
    descendants = rows[1:]
    date_values = [value for value in [feature_row["start_date"], feature_row["due_date"]] if pd.notnull(value)]
    for row in descendants:
        if pd.notnull(row["start_date"]):
            date_values.append(row["start_date"])
        if pd.notnull(row["due_date"]):
            date_values.append(row["due_date"])
    if not date_values:
        return

    range_start = monday_floor(min(date_values))
    range_end = monday_floor(max(date_values))
    weeks = pd.date_range(start=range_start, end=range_end, freq="W-MON")
    if len(weeks) == 0:
        weeks = pd.DatetimeIndex([range_start])
    week_ranges = [(week, week + timedelta(days=6)) for week in weeks]

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4 + len(week_ranges))
    title_cell = ws.cell(1, 1)
    title_cell.value = safe_excel_text(title)
    title_cell.font = Font(bold=True, size=14, color=theme.text_muted)
    title_cell.alignment = Alignment(horizontal="left", vertical="center")

    month_groups = []
    current_month = None
    current_start = None
    for offset, (week_start, _) in enumerate(week_ranges, start=5):
        label = week_start.strftime("%B %Y")
        if label != current_month:
            if current_month is not None:
                month_groups.append((current_start, offset - 1, current_month))
            current_month = label
            current_start = offset
    if current_month is not None:
        month_groups.append((current_start, 4 + len(week_ranges), current_month))

    for start_col, end_col, label in month_groups:
        ws.merge_cells(start_row=2, start_column=start_col, end_row=2, end_column=end_col)
        cell = ws.cell(2, start_col)
        cell.value = label
        cell.fill = solid(theme.header_dark)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for col in range(1, 5):
        ws.cell(2, col).fill = solid(theme.header_dark)
        ws.cell(2, col).border = make_border(theme)

    headers = ["Stage", "Activity", "Start", "Due"] + [f"W{week.isocalendar().week} {week.strftime('%d/%m/%y')}" for week, _ in week_ranges]
    ws.append(headers)
    for cell in ws[3]:
        cell.fill = solid(theme.header_mid)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = make_border(theme)

    status_map = build_status_map(pd.DataFrame(rows), theme)
    feature_values = ["Feature", safe_excel_text(feature_row["task_name"]), format_cell_date(feature_row["start_date"], True), format_cell_date(feature_row["due_date"], True)]
    ws.append(feature_values)
    feature_row_idx = ws.max_row
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(feature_row_idx, col)
        cell.fill = solid(theme.header_light)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = make_border(theme)
    for col, (week_start, week_end) in enumerate(week_ranges, start=5):
        if feature_row["start_date"] <= week_end and feature_row["due_date"] >= week_start:
            ws.cell(feature_row_idx, col).fill = solid(theme.parent_bar)

    band_toggle = False
    for row in descendants:
        stage, activity = split_stage_activity(row["task_name"])
        values = [safe_excel_text(stage), safe_excel_text(activity), format_cell_date(row["start_date"], True), format_cell_date(row["due_date"], True)]
        ws.append(values)
        row_idx = ws.max_row

        fill = solid(theme.milestone_row) if stage == "Milestone" else solid(theme.task_alt if band_toggle else theme.task_plain)
        if stage != "Milestone":
            band_toggle = not band_toggle
        for col in range(1, 5):
            cell = ws.cell(row_idx, col)
            cell.fill = fill
            cell.border = make_border(theme)
            cell.alignment = Alignment(vertical="center", wrap_text=True)

        colour = theme.milestone_bar if stage == "Milestone" else bar_colour(row, colour_by, theme, status_map)
        if pd.notnull(row["start_date"]) and pd.notnull(row["due_date"]):
            for col, (week_start, week_end) in enumerate(week_ranges, start=5):
                if row["start_date"] <= week_end and row["due_date"] >= week_start:
                    ws.cell(row_idx, col).fill = solid(colour)
                    ws.cell(row_idx, col).border = make_border(theme)

    widths = [22, 44, 13, 13]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    for col in range(5, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col)].width = 11


def build_date_groups(days: pd.DatetimeIndex, start_col: int) -> tuple[list[tuple[int, int, str]], list[tuple[int, int, str]]]:
    month_groups = []
    current_month = None
    current_start = None
    for offset, day in enumerate(days, start=start_col):
        label = day.strftime("%B %Y")
        if label != current_month:
            if current_month is not None:
                month_groups.append((current_start, offset - 1, current_month))
            current_month = label
            current_start = offset
    if current_month is not None:
        month_groups.append((current_start, start_col - 1 + len(days), current_month))

    week_groups = []
    current_week = None
    current_week_start = None
    for offset, day in enumerate(days, start=start_col):
        label = f"W{day.isocalendar().week} {day.year}"
        if label != current_week:
            if current_week is not None:
                week_groups.append((current_week_start, offset - 1, current_week))
            current_week = label
            current_week_start = offset
    if current_week is not None:
        week_groups.append((current_week_start, start_col - 1 + len(days), current_week))
    return month_groups, week_groups


def build_week_groups(week_ranges: list[tuple[pd.Timestamp, pd.Timestamp]], start_col: int) -> list[tuple[int, int, str]]:
    month_groups = []
    current_month = None
    current_start = None
    for offset, (week_start, _) in enumerate(week_ranges, start=start_col):
        label = week_start.strftime("%B %Y")
        if label != current_month:
            if current_month is not None:
                month_groups.append((current_start, offset - 1, current_month))
            current_month = label
            current_start = offset
    if current_month is not None:
        month_groups.append((current_start, start_col - 1 + len(week_ranges), current_month))
    return month_groups


def build_workbook(frame: pd.DataFrame, config: dict[str, Any]) -> Workbook:
    start_timeline, end_timeline = infer_timeline(frame, config)
    ordered = rebuild_hierarchy(frame)
    descendant_ranges = compute_descendant_ranges(ordered)
    output_cfg = config["output"]
    theme = build_theme(config)
    wb = Workbook()

    modes = set(output_cfg.get("modes", ["daily", "weekly"]))
    omit_owner = output_cfg.get("omit_owner", False)
    colour_by = output_cfg.get("colour_by", "owner")
    ignore_parent_bars = output_cfg.get("ignore_parent_bars", False)

    if "daily" in modes:
        days = pd.date_range(start=start_timeline, end=end_timeline, freq="D")
        start_col = 4 if omit_owner else 5
        month_groups, week_groups = build_date_groups(days, start_col)
        write_main_sheet(
            wb,
            "Gantt - Daily",
            ordered,
            list(days),
            [day.strftime("%d/%m/%y") for day in days],
            month_groups,
            week_groups,
            colour_by,
            omit_owner,
            ignore_parent_bars,
            descendant_ranges,
            theme,
        )

    if "weekly" in modes:
        weeks = pd.date_range(start=start_timeline, end=end_timeline, freq="W-MON")
        if len(weeks) == 0:
            weeks = pd.DatetimeIndex([monday_floor(start_timeline)])
        week_ranges = [(week, week + timedelta(days=6)) for week in weeks]
        start_col = 4 if omit_owner else 5
        month_groups = build_week_groups(week_ranges, start_col)
        write_main_sheet(
            wb,
            "Gantt - Weekly",
            ordered,
            week_ranges,
            [f"W{week.isocalendar().week} {week.strftime('%d/%m/%y')}" for week, _ in week_ranges],
            month_groups,
            [],
            colour_by,
            omit_owner,
            ignore_parent_bars,
            descendant_ranges,
            theme,
        )

    if output_cfg.get("focus_sheets", False):
        for candidate in focus_sheet_candidates(ordered):
            rows = descendants_for(ordered, candidate["task_id"])
            if len(rows) > 1:
                build_focus_sheet(wb, candidate["task_name"], rows, colour_by, theme)

    if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
        del wb["Sheet"]
    return wb


def main() -> int:
    args = parse_args()
    config = merge_dict(DEFAULT_CONFIG, load_json(Path(args.config)))
    config = apply_overrides(config, args)
    frame = normalize_frame(read_input(config), config)
    workbook = build_workbook(frame, config)
    output_path = Path(config["output"]["path"])
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    print(f"Gantt generated: {output_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
